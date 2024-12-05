from django.core.management.base import BaseCommand
from pathlib import Path
from indicator.models import Indicator
from institution.models import Institution
import csv
from django.conf import settings
import openpyxl
import pandas as pd
import warnings
warnings.filterwarnings("ignore")
REL_PATH_INDICATORS = 'install/postgres/indicators.xlsx'


class Command(BaseCommand):
    help = """Import indicators from a CSV file into the Indicators model.
    Only used during installation of the application."""

    def handle(self, *args, **kwargs):
        input_file = Path(settings.BASE_DIR.parent.resolve() /
                          REL_PATH_INDICATORS)
        if not input_file.exists():
            self.stderr.write(f"File {input_file} not found.")
            return
        self.stdout.write(f'Importing indicators from {input_file}...')
        workbook = openpyxl.load_workbook(input_file)
        first_sheet = workbook.sheetnames[0]
        sheet = workbook[first_sheet]
        for row in sheet.iter_rows(values_only=True, min_row=2):
            institution, name, group, description, unit, abbreviation = row
            inst_instid = Institution.objects.get(abbreviation=institution)
            created_at = pd.to_datetime('now')
            updated_at = pd.to_datetime('now')

            if Indicator.objects.filter(inst_instid=inst_instid,
                                        abbreviation=abbreviation).exists():
                self.stderr.write(
                    f'Indicator {name} already exists. Skipping...')
                continue

            Indicator.objects.create(
                inst_instid=inst_instid,
                name=name,
                group=group,
                description=description,
                unit=unit,
                created_at=created_at,
                updated_at=updated_at,
                abbreviation=abbreviation
            )
        self.stdout.write(self.style.SUCCESS(
            'Indicators imported successfully.'))
